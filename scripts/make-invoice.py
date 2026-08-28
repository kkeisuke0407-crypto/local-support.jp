# -*- coding: utf-8 -*-
"""A4縦1枚にちょうど収まる請求書を作る。
   A4(210x297) − 左右余白12.7mm×2 − 上下余白12.7mm×2 → 使用領域 約184.6 x 271.6mm
   Excel列幅 w → px = w*7+5 、px→mm = *25.4/96
   行高は pt 指定、pt→mm = *0.3528
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

F='Arial'
def f(sz=10,b=False,c='FF1A1A1A'): return Font(name=F,size=sz,bold=b,color=c)
thin=Side('thin',color='FF8C8C8C'); med=Side('medium',color='FF1A1A1A'); hair=Side('hair',color='FFCCCCCC')
YEL=PatternFill('solid',fgColor='FFFFFF00'); GRY=PatternFill('solid',fgColor='FFEFEFEF')
LGY=PatternFill('solid',fgColor='FFF7F7F7'); DGY=PatternFill('solid',fgColor='FFE2E2E2')
L=Alignment('left',vertical='center'); LT=Alignment('left',vertical='top',wrap_text=True)
LC=Alignment('left',vertical='center',wrap_text=True)
R=Alignment('right',vertical='center'); C=Alignment('center',vertical='center')

wb=Workbook(); ws=wb.active; ws.title='請求書'

# ── 列幅：合計 約185mm ──────────────────────────
W={'A':1.6,'B':30,'C':7,'D':12,'E':13,'F':1.6,'G':11,'H':19}
for k,v in W.items(): ws.column_dimensions[k].width=v
px=sum(v*7+5 for v in W.values()); print(f'横幅 {px:.0f}px = {px*25.4/96:.1f}mm')

def P(cell,val=None,font=None,align=None,fill=None,border=None,fmt=None):
    c=ws[cell]
    if val is not None: c.value=val
    if font: c.font=font
    if align: c.alignment=align
    if fill: c.fill=fill
    if border: c.border=border
    if fmt: c.number_format=fmt
    return c

def box(r1,c1,r2,c2,side=thin):
    """矩形に外枠を引く"""
    cols=[chr(ord(c1)+i) for i in range(ord(c2)-ord(c1)+1)]
    for r in range(r1,r2+1):
        for col in cols:
            b=ws[f'{col}{r}'].border
            ws[f'{col}{r}'].border=Border(
                left = side if col==c1 else b.left,
                right= side if col==c2 else b.right,
                top  = side if r==r1  else b.top,
                bottom=side if r==r2  else b.bottom)

H={}   # 行高(pt)

# ── タイトル ─────────────────────────────
H[1]=6
ws.merge_cells('B2:H2'); P('B2','請　求　書',f(24,True),C); H[2]=42
for col in 'BCDEFGH': ws[f'{col}2'].border=Border(bottom=med)
H[3]=6

# ── 上部：宛名（左） / 発行者（右）────────────────
P('G4','請求書番号',f(9,c='FF666666'),R); P('H4','LS-2026-0828-001',f(9),R); H[4]=16
ws.merge_cells('B5:E5'); P('B5','竹内産業株式会社 エネルギー事業部 御中',f(14,True),L)
for col in 'BCDE': ws[f'{col}5'].border=Border(bottom=Side('thin',color='FF1A1A1A'))
P('G5','発行日',f(9,c='FF666666'),R); P('H5','2026年8月28日',f(9),R); H[5]=28
H[6]=10
ws.merge_cells('B7:E7'); P('B7','下記のとおりご請求申し上げます。',f(10),L)
P('G7','ローカル情報局',f(12,True),R); H[7]=20
for i,(lb,vl) in enumerate([('事業主','藤波 圭佑'),('登録番号','T2810572974271'),
                            ('所在地',None),('連絡先','support@local-support.jp')]):
    r=8+i; H[r]=17
    P(f'G{r}',lb,f(9,c='FF666666'),R)
    if vl is None: P(f'H{r}',None,f(9),R,fill=YEL,border=Border(bottom=hair))
    else:          P(f'H{r}',vl,f(9),R)
H[12]=10

# ── ご請求金額 ────────────────────────────
ws.merge_cells('B13:C14'); P('B13','ご請求金額',f(12,True),C)
ws.merge_cells('D13:E14'); P('D13',None,f(22,True),C,fmt='¥#,##0')
box(13,'B',14,'E',med); H[13]=22; H[14]=22
P('F14','（消費税込）',f(9,c='FF666666'),L)
H[15]=10

# ── 明細 ─────────────────────────────────
HR=16
for col,lb,al in [('B','品目',C),('C','数量',C),('D','単価',C),('E','金額',C)]:
    P(f'{col}{HR}',lb,f(10,True),al,fill=GRY,border=Border(thin,thin,thin,thin))
H[HR]=22

FIRST=HR+1; ROWS=8; LAST=HR+ROWS
P(f'B{FIRST}','業者ご紹介手数料\n案件ID：SOL-2026-0718-001\n太陽光パネル清掃・除草（兵庫県佐用町）／成約金額 88,000円（税込）の10%',f(10),LT)
P(f'C{FIRST}','1式',f(10),C)
P(f'D{FIRST}',8000,f(10),R,fmt='#,##0')
P(f'E{FIRST}',f'=IF(D{FIRST}="","",D{FIRST})',f(10),R,fmt='#,##0')
H[FIRST]=52
for r in range(FIRST+1,LAST+1):
    P(f'E{r}',f'=IF(D{r}="","",D{r})',f(10),R,fmt='#,##0')
    H[r]=21
for r in range(FIRST,LAST+1):
    for col in 'BCDE':
        ws[f'{col}{r}'].border=Border(thin,thin,thin,thin)
        if col=='B' and r>FIRST: ws[f'{col}{r}'].alignment=LC
H[LAST+1]=8

# ── 集計 ─────────────────────────────────
S=LAST+2
P(f'D{S}',  '小計（10%対象）',f(10),L,fill=LGY,border=Border(thin,thin,thin,thin))
P(f'E{S}',  f'=SUM(E{FIRST}:E{LAST})',f(10),R,border=Border(thin,thin,thin,thin),fmt='#,##0" 円"')
P(f'D{S+1}','消費税（10%）',  f(10),L,fill=LGY,border=Border(thin,thin,thin,thin))
P(f'E{S+1}',f'=ROUND(E{S}*0.1,0)',f(10),R,border=Border(thin,thin,thin,thin),fmt='#,##0" 円"')
P(f'D{S+2}','合計',f(11,True),L,fill=DGY,border=Border(thin,thin,thin,thin))
P(f'E{S+2}',f'=E{S}+E{S+1}',f(11,True),R,fill=DGY,border=Border(thin,thin,thin,thin),fmt='#,##0" 円"')
for i in range(3): H[S+i]=21
TOTAL=S+2
ws['D13']='=E%d'%TOTAL   # 合計行が確定してから参照を書く
H[TOTAL+1]=12

# ── お振込先 ──────────────────────────────
PR=TOTAL+2
ws.merge_cells(f'B{PR}:E{PR}'); P(f'B{PR}','お振込先',f(10,True),L,fill=LGY); H[PR]=20
for i,lb in enumerate(['金融機関','種別・番号','口座名義']):
    r=PR+1+i; H[r]=20
    P(f'B{r}',f'　{lb}',f(10,c='FF555555'),L)
    ws.merge_cells(f'C{r}:E{r}')
    P(f'C{r}',None,f(10),L,fill=YEL,border=Border(bottom=hair))
PEND=PR+3
box(PR,'B',PEND,'E')
H[PEND+1]=12

# ── 備考 ─────────────────────────────────
BR=PEND+2
ws.merge_cells(f'B{BR}:H{BR}'); P(f'B{BR}','備考',f(10,True),L,fill=LGY); H[BR]=20
notes=['お支払期限：2026年9月30日（8月末締め・翌月末払い）',
       '振込手数料は恐れ入りますが貴社にてご負担をお願いいたします。',
       '本請求書は適格請求書（インボイス）として発行しております。',
       '「ロカサポ（local-support.jp）」は屋号「ローカル情報局」として藤波が運営しております。',
       '日頃のご連絡は運営担当の寺尾より差し上げております。']
for i,n in enumerate(notes):
    r=BR+1+i; H[r]=17
    ws.merge_cells(f'B{r}:H{r}')
    P(f'B{r}',f'　・{n}' if i<4 else f'　　{n}',f(9,c='FF333333'),L)
BEND=BR+len(notes)
box(BR,'B',BEND,'H')

for r,h in H.items(): ws.row_dimensions[r].height=h
total_pt=sum(H.values()); print(f'縦の高さ {total_pt}pt = {total_pt*0.3528:.1f}mm（A4使用領域 約271mm）')

# ── 凡例（印刷範囲外）─────────────────────────
LG=BEND+3
P(f'B{LG}','【記入方法 — 印刷範囲の外なのでPDFには出ません】',f(10,True,'FFC00000'),L)
P(f'B{LG+1}',f'黄色いセルに入力してください：H10（所在地）／C{PR+1}（金融機関）／C{PR+2}（種別・番号）／C{PR+3}（口座名義）',f(9,c='FFC00000'),L)
P(f'B{LG+2}','入力例：〒150-0002 東京都渋谷区渋谷1-2-3 ／ ◯◯銀行 △△支店 ／ 普通 1234567 ／ フジナミ ケイスケ',f(9,c='FF777777'),L)
P(f'B{LG+3}',f'金額を変えるときは単価（D{FIRST}）だけ直せば、金額・小計・消費税・合計・上部の請求金額まで自動で再計算されます。',f(9,c='FFC00000'),L)
P(f'B{LG+4}',f'明細を追加するときは {FIRST+1}〜{LAST} 行目の単価（D列）に入力してください。金額欄は自動で入ります。',f(9,c='FFC00000'),L)

# ── ページ設定 ────────────────────────────
ws.print_area=f'A1:H{BEND}'
ws.page_setup.orientation='portrait'
ws.page_setup.paperSize=ws.PAPERSIZE_A4
ws.sheet_properties.pageSetUpPr.fitToPage=True
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
ws.page_margins.left=ws.page_margins.right=0.5
ws.page_margins.top=ws.page_margins.bottom=0.5
ws.page_margins.header=ws.page_margins.footer=0
ws.sheet_view.showGridLines=False
ws.print_options.horizontalCentered=True

wb.save('docs/invoice/SOL-2026-0718-001.xlsx')
print(f'明細 {FIRST}-{LAST} / 集計 {S}-{TOTAL} / 振込先 {PR}-{PEND} / 備考 {BR}-{BEND} / 印刷範囲 A1:H{BEND}')
